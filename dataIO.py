import math
import openpyxl
import os
from Objects.Line import Line
from Objects.DbText import DbText
from tkinter import messagebox, filedialog


def get_data(file_name: str, dir_path: str):
    abs_file_path: str = dir_path + '/' + file_name
    data = [{}, []]
    key_coordinate: tuple[float, float]  # for detect coordinate with high accuracy
    with open(abs_file_path, 'r', encoding='utf-8') as f:
        for obj in f:
            # duplicate check
            if str(obj).startswith("$ListLine"):  # if line
                spl = obj.split("||")
                if math.isclose(float(spl[6]), 0, rel_tol=1e-1):  # if horizontal
                    key_coordinate = (round((float(spl[1]) + float(spl[2])) / 2), round(float(spl[3])))
                else:  # if vertical
                    key_coordinate = (round((float(spl[3]) + float(spl[4])) / 2), round(float(spl[1])))
                for i in range(-1, 2):
                    for j in range(-1, 2):
                        exist = data[0].get((key_coordinate[0] + i, key_coordinate[1] + j))
                        if exist:
                            break
                    if exist:
                        break
                if exist:
                    continue
                #

                line = Line(obj.split('||'), file_name)
                if line.vertical is not None:
                    data[0][key_coordinate] = line

            else:  # if text
                data[1].append(DbText(obj.split("||")))
    return data


def write_pf_to_excel(pf_list: list):
    def insert_into_cells(sheet, row: int, columns: dict):
        for col, value in columns.items():
            sheet.cell(row=row, column=col).value = value

    #  init excel file
    selected = False
    while selected is False:
        file_path = filedialog.askopenfilename(title="Выберите файл excel куда нужно записать данные")
        if file_path == "":
            exit()
        if os.path.splitext(file_path)[1] in (".xlsx", "xls"):
            selected = True
        else:
            messagebox.showwarning("Внимание!", "Расширение файла должно быть .xlsx или .xls")

    wb = openpyxl.load_workbook(file_path)
    pf_sheet = wb["Техпроцессы"]
    operations_sheet = wb["Операции"]
    shifts_sheet = wb["Переходы"]
    ammo_sheet = wb["Оснаcтка"]
    unresolved_sheet = wb["Нераспознано"]
    kk_sheet = wb["КК"]
    #

    #  check already existed process flows
    already_existing_in_excel = {}
    already_existing_match = {}
    for r in range(2, pf_sheet.max_row + 1):
        current_cell_value = pf_sheet.cell(row=r, column=2).value
        if already_existing_in_excel.get(current_cell_value) is None:
            already_existing_in_excel[current_cell_value] = True
    #

    act_row_pf_sheet = pf_sheet.max_row + 1
    act_row_operation_sheet = operations_sheet.max_row + 1
    act_row_shift_sheet = shifts_sheet.max_row + 1
    act_row_ammo_sheet = ammo_sheet.max_row + 1
    act_row_unresolved_sheet = unresolved_sheet.max_row + 1

    # Insert MAIN (process flow) info in main sheet
    for pf in pf_list:
        if already_existing_in_excel.get(pf.marker) is not None:
            already_existing_match[pf.marker] = pf.file_name
            continue

        insert_into_cells(pf_sheet, act_row_pf_sheet, {1: pf.name,
                                                       2: pf.marker,
                                                       3: pf.description,
                                                       4: pf.checker,
                                                       5: pf.checker,
                                                       6: pf.approver,
                                                       8: pf.normalizator,
                                                       9: pf.liter,
                                                       }
                          )
        act_row_pf_sheet += 1
        #

        # Insert UNRESOLVED info in main sheet
        for unresolved in pf.unresolved_text:
            insert_into_cells(unresolved_sheet, act_row_unresolved_sheet, {1: pf.marker,
                                                                           2: unresolved.number_of_operation,
                                                                           3: unresolved.text,
                                                                           }
                              )
            act_row_unresolved_sheet += 1
        #

        # Insert OPERATION info in main sheet
        for operation in dict(pf.operations).values():
            insert_into_cells(operations_sheet, act_row_operation_sheet, {1: pf.marker,
                                                                          3: operation.number,
                                                                          4: operation.description,
                                                                          5: operation.ammo,
                                                                          7: operation.IOT,
                                                                          9: operation.tsekh,
                                                                          }
                              )
            act_row_operation_sheet += 1
            #

            # Insert SHIFT info in main sheet
            for shift in operation.shifts:
                insert_into_cells(shifts_sheet, act_row_shift_sheet, {1: pf.marker,
                                                                      2: operation.number,
                                                                      3: shift.number,
                                                                      4: shift.description,
                                                                      5: shift.percent_control,
                                                                      }
                                  )
                act_row_shift_sheet += 1
                #

                # Insert TOOL info in main sheet
                for tool in shift.tools:
                    insert_into_cells(ammo_sheet, act_row_ammo_sheet, {1: pf.marker,
                                                                       2: operation.number,
                                                                       3: shift.number,
                                                                       5: tool.name,
                                                                       }
                                      )
                    act_row_ammo_sheet += 1
                #

    wb.save(file_path)
    messagebox.showinfo("Успешно!", f"Из {len(pf_list)} добавлено в файл {len(pf_list) - len(already_existing_match)}\n"
                                    f"Если что-то не добавлено, значит такой ТП уже был в файле.\n\n"
                                    f"Список недобаленных ТП:\n\n"
                                    f"{list(already_existing_match.values())}")
