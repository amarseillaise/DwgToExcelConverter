import os
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
from Objects.ProcessFlow import ProcessFlow


def write_pf_to_excel(pf_list):
    #  init excel file
    selected = False
    while selected is False:
        file_path = filedialog.askopenfilename(title="Выберите файл excel куда нужно записать данные")
        if file_path == "":
            exit()
        if os.path.splitext(file_path)[1] in (".xlsx", "xls"):
            selected = True
        else:
            messagebox.showwarning("Внимание!", "Расширение файла должно быть .xlsx или .xls")

    wb = openpyxl.load_workbook(file_path)
    pf_sheet = wb["Техпроцессы"]
    operations_sheet = wb["Операции"]
    shifts_sheet = wb["Переходы"]
    ammo_sheet = wb["Оснаcтка"]
    unresolved_sheet = wb["Нераспознано"]
    kk_sheet = wb["КК"]
    #

    #  check already existed process flows
    already_existing_in_excel = {}
    already_existing_match = {}
    for r in range(2, pf_sheet.max_row + 1):
        current_cell_value = pf_sheet.cell(row=r, column=2).value
        if already_existing_in_excel.get(current_cell_value) is None:
            already_existing_in_excel[current_cell_value] = True
    #

    act_row_pf_sheet = pf_sheet.max_row + 1
    act_row_operation_sheet = operations_sheet.max_row + 1
    act_row_shift_sheet = shifts_sheet.max_row + 1
    act_row_ammo_sheet = ammo_sheet.max_row + 1
    act_row_unresolved_sheet = unresolved_sheet.max_row + 1

    # Insert MAIN (process flow) info in main sheet
    for pf in pf_list:
        if already_existing_in_excel.get(pf.marker) is not None:
            already_existing_match[pf.pf.marker] = pf.__file_name
            continue

        pf_sheet.cell(row=act_row_pf_sheet, column=1).value = pf.name
        pf_sheet.cell(row=act_row_pf_sheet, column=2).value = pf.marker
        pf_sheet.cell(row=act_row_pf_sheet, column=3).value = pf.description
        pf_sheet.cell(row=act_row_pf_sheet, column=4).value = pf.developer
        pf_sheet.cell(row=act_row_pf_sheet, column=5).value = pf.checker
        pf_sheet.cell(row=act_row_pf_sheet, column=6).value = pf.approver
        pf_sheet.cell(row=act_row_pf_sheet, column=8).value = pf.normalizator
        pf_sheet.cell(row=act_row_pf_sheet, column=9).value = pf.liter

        act_row_pf_sheet += 1
    #

        # Insert UNRESOLVED info in main sheet
        for unresolved in pf.unresolved_text:
            unresolved_sheet.cell(row=act_row_unresolved_sheet, column=1).value = pf.marker
            unresolved_sheet.cell(row=act_row_unresolved_sheet, column=2).value = unresolved.number_of_operation
            unresolved_sheet.cell(row=act_row_unresolved_sheet, column=3).value = unresolved.text

            act_row_unresolved_sheet += 1
        #


        # Insert OPERATION info in main sheet
        for operation in dict(pf.operations).values():
            operations_sheet.cell(row=act_row_operation_sheet, column=1).value = pf.marker
            operations_sheet.cell(row=act_row_operation_sheet, column=3).value = operation.number
            operations_sheet.cell(row=act_row_operation_sheet, column=4).value = operation.description
            operations_sheet.cell(row=act_row_operation_sheet, column=5).value = operation.ammo
            operations_sheet.cell(row=act_row_operation_sheet, column=7).value = operation.IOT
            operations_sheet.cell(row=act_row_operation_sheet, column=9).value = operation.tsekh

            act_row_operation_sheet += 1
        #

            # Insert SHIFT info in main sheet
            for shift in operation.shifts:
                shifts_sheet.cell(row=act_row_shift_sheet, column=1).value = pf.marker
                shifts_sheet.cell(row=act_row_shift_sheet, column=2).value = operation.number
                shifts_sheet.cell(row=act_row_shift_sheet, column=3).value = shift.number
                shifts_sheet.cell(row=act_row_shift_sheet, column=4).value = shift.description
                shifts_sheet.cell(row=act_row_shift_sheet, column=5).value = shift.percent_control

                act_row_shift_sheet += 1
            #

                # Insert TOOL info in main sheet
                for tool in shift.tools:
                    ammo_sheet.cell(row=act_row_ammo_sheet, column=1).value = pf.marker
                    ammo_sheet.cell(row=act_row_ammo_sheet, column=2).value = operation.number
                    ammo_sheet.cell(row=act_row_ammo_sheet, column=3).value = shift.number
                    ammo_sheet.cell(row=act_row_ammo_sheet, column=5).value = tool.name

                    act_row_ammo_sheet += 1
                #

    wb.save(file_path)
    messagebox.showinfo("Успешно!", f"Из {len(pf_list)} добавлено в файл {len(pf_list) - len(already_existing_match)}\n"
                                    f"Если что-то не добавлено, значит такой ТП уже был в файле.\n\n"
                                    f"Список недобаленных ТП:\n\n"
                                    f"{already_existing_match.values()}")





